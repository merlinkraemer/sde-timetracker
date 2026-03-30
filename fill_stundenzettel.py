#!/usr/bin/env python3
"""
fill_stundenzettel.py - Fill XLSX timesheet from clawd time tracker

Reads time entries from clawd/sde-time-tracker.md and adds them to the
Stundenzettel XLSX.

Usage:
    .venv/bin/python3 fill_stundenzettel.py                    # Fill last complete month
    .venv/bin/python3 fill_stundenzettel.py --month 2026-03    # Fill specific month
    .venv/bin/python3 fill_stundenzettel.py --dry-run           # Preview without writing

Setup (one-time):
    python3 -m venv .venv && .venv/bin/pip install openpyxl
"""

import re
import sys
import datetime
import calendar
import argparse
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl not installed.")
    print("Run: python3 -m venv .venv && .venv/bin/pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = Path(__file__).parent
CLAWD_MD = SCRIPT_DIR / "clawd" / "sde-time-tracker.md"

# German month abbreviations for tab names
DE_MONTH_ABBR = {
    1: "Jan", 2: "Feb", 3: "Mär", 4: "Apr", 5: "Mai", 6: "Jun",
    7: "Jul", 8: "Aug", 9: "Sep", 10: "Okt", 11: "Nov", 12: "Dez",
}

# Month name -> number (English + German, for parsing markdown headers)
MONTH_NAME_TO_NUM = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
    "januar": 1, "februar": 2, "märz": 3, "april": 4,
    "mai": 5, "juni": 6, "juli": 7, "august": 8,
    "september": 9, "oktober": 10, "november": 11, "dezember": 12,
}

# Shared styles
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
FONT_10 = Font(size=10)
FONT_10B = Font(size=10, bold=True)
FONT_8B = Font(size=8, bold=True)

# Layout constant (new format, Jan 2026+)
DATA_START_ROW = 6      # first date row


def tab_name(year: int, month: int) -> str:
    return f"{DE_MONTH_ABBR[month]}{year % 100:02d}"


def find_latest_xlsx() -> Path:
    candidates = sorted(
        SCRIPT_DIR.glob("*/Stundenzettel-*-Merlin.xlsx"),
        key=lambda p: p.stat().st_mtime,
    )
    if not candidates:
        print("ERROR: No Stundenzettel-*-Merlin.xlsx found under", SCRIPT_DIR)
        sys.exit(1)
    return candidates[-1]


def parse_time(s: str):
    s = s.strip()
    m = re.match(r"(\d{1,2}):(\d{2})", s)
    if not m:
        return None
    return datetime.time(int(m.group(1)), int(m.group(2)))


def parse_month_entries(md_text: str, year: int, month: int) -> dict:
    """Return {date: [{start, end, notes}, ...]} for completed sessions.

    Multiple sessions on the same day are kept as separate entries
    so they get separate rows in the XLSX (matching manual workflow).
    """

    # Build regex that matches English or German month name for this month
    names = [n for n, num in MONTH_NAME_TO_NUM.items() if num == month]
    name_pattern = "|".join(re.escape(n) for n in names)
    header_re = re.compile(
        rf"^##\s+({name_pattern})\s+{year}\s*$", re.IGNORECASE | re.MULTILINE
    )
    match = header_re.search(md_text)
    if not match:
        print(f"ERROR: No section header for month {month}/{year} found in markdown")
        print(f"  Looked for pattern: ## ({name_pattern}) {year}")
        sys.exit(1)

    section_start = match.end()
    next_header = re.search(r"^## ", md_text[section_start:], re.MULTILINE)
    section = md_text[section_start : section_start + next_header.start()] if next_header else md_text[section_start:]

    entries: dict[datetime.date, list] = {}
    for line in section.splitlines():
        line = line.strip()
        if not line.startswith("|") or "Date" in line or line.startswith("|--"):
            continue
        cols = [c.strip() for c in line.split("|")]
        cols = [c for c in cols if c != ""]  # drop empties from leading/trailing |
        if len(cols) < 4:
            continue

        date_str, start_str, end_str, hours_str = cols[0], cols[1], cols[2], cols[3]
        note = cols[4] if len(cols) > 4 else ""

        # Skip incomplete or cancelled sessions
        if end_str in ("-", "") or hours_str in ("-", ""):
            print(f"  Skipping incomplete: {date_str}")
            continue

        try:
            date = datetime.date.fromisoformat(date_str)
        except ValueError:
            continue

        start_t = parse_time(start_str)
        end_t = parse_time(end_str)
        if start_t is None or end_t is None:
            continue

        entries.setdefault(date, []).append(
            {"start": start_t, "end": end_t, "notes": note}
        )

    return entries


def _style_cell(cell, font=FONT_10, nf=None, align=None, border=THIN_BORDER):
    cell.font = font
    if nf:
        cell.number_format = nf
    if align:
        cell.alignment = Alignment(horizontal=align)
    if border:
        cell.border = border


def create_month_sheet(wb, year: int, month: int, entries: dict) -> tuple[str, int]:
    """Create/overwrite a month tab.

    Returns (tab_name, überstunden_cell_row) so the caller knows which
    cell to reference from the Überstunden tab.
    """
    name = tab_name(year, month)
    if name in wb.sheetnames:
        print(f"  Tab '{name}' exists — overwriting data rows")
        ws = wb[name]
    else:
        ub_idx = wb.sheetnames.index("Überstunden") if "Überstunden" in wb.sheetnames else len(wb.sheetnames)
        ws = wb.create_sheet(name, ub_idx)

    days = calendar.monthrange(year, month)[1]

    # Column widths
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 17.4
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["F"].width = 40.4

    # Row 1
    _style_cell(ws["A1"], font=FONT_10)
    ws["A1"] = "Aufwandsnachweis für BV. "
    _style_cell(ws["F1"], font=FONT_10B)
    ws["F1"] = "SUB-events GmbH"

    # Row 2
    ws["A2"] = "NAME"
    ws["B2"] = "Merlin Krämer"
    _style_cell(ws["F2"], font=FONT_10B)
    ws["F2"] = "Lauenbruch Ost 1, 21079 Hamburg"

    # Row 4
    ws["B4"] = "incl. Pausen + Fahrtzeit"

    # Row 5 — column headers
    for col_letter, text in [("A", "Datum"), ("B", "von"), ("C", "bis"),
                              ("D", "Std."), ("E", "Ort / Projekt"), ("F", "Erläuterungen")]:
        c = ws[f"{col_letter}5"]
        c.value = text
        _style_cell(c, font=FONT_8B)

    # Data rows — one per calendar day, extra rows for multi-session days
    row = DATA_START_ROW
    session_count = 0
    for day in range(1, days + 1):
        date = datetime.date(year, month, day)
        day_entries = entries.get(date, [])

        # Always emit at least one row per calendar day (even if no sessions)
        rows_for_day = max(1, len(day_entries))
        for i in range(rows_for_day):
            a = ws.cell(row=row, column=1, value=datetime.datetime(year, month, day))
            _style_cell(a, nf="mm-dd-yy", align="center")

            b = ws.cell(row=row, column=2)
            _style_cell(b, nf="h:mm")

            c_ = ws.cell(row=row, column=3)
            _style_cell(c_, nf="h:mm")

            d = ws.cell(row=row, column=4, value=f"=(C{row}-B{row}+(C{row}<B{row}))*24")
            _style_cell(d, nf="0.00")

            e = ws.cell(row=row, column=5)
            _style_cell(e, align="left")

            f = ws.cell(row=row, column=6)
            _style_cell(f)

            if i < len(day_entries):
                entry = day_entries[i]
                b.value = entry["start"]
                c_.value = entry["end"]
                if entry["notes"]:
                    f.value = entry["notes"]
                session_count += 1

            row += 1

    last_data_row = row - 1

    # Summary rows: 2 empty rows gap after data, then Gesamt, skip 1, Max, Überstunden
    gesamt_row = last_data_row + 3
    max_row = gesamt_row + 2
    ub_row = max_row + 1

    # Gesamt
    a = ws.cell(row=gesamt_row, column=1, value="Gesamt")
    _style_cell(a, font=FONT_10B, align="right")
    for col in (2, 3):
        _style_cell(ws.cell(row=gesamt_row, column=col), font=FONT_10B, align="right")
    d = ws.cell(row=gesamt_row, column=4, value=f"=SUM(D{DATA_START_ROW}:D{last_data_row})")
    _style_cell(d, font=FONT_10B, nf="0.00", align="center")

    # Max Std
    a = ws.cell(row=max_row, column=1, value="Max. Std. Auszahlung (556\u20ac/18,50\u20ac Stundenlohn)")
    _style_cell(a)
    for col in (2, 3):
        _style_cell(ws.cell(row=max_row, column=col))
    d = ws.cell(row=max_row, column=4, value="=556/18.5")
    _style_cell(d, nf="0.00", align="center")

    # Überstunden
    a = ws.cell(row=ub_row, column=1, value="Überstunden")
    _style_cell(a, align="right")
    for col in (2, 3):
        _style_cell(ws.cell(row=ub_row, column=col))
    d = ws.cell(row=ub_row, column=4, value=f"=D{gesamt_row}-D{max_row}")
    _style_cell(d, nf="0.00", align="center")

    extra = sum(max(0, len(v) - 1) for v in entries.values())
    print(f"  Created tab '{name}': {days} days, {session_count} sessions"
          + (f" ({extra} extra rows for multi-session days)" if extra else ""))
    return name, ub_row


def update_ueberstunden_tab(wb, month_tab: str, year: int, month: int, ub_cell_row: int):
    """Add/update the month in the Überstunden summary tab."""
    ws = wb["Überstunden"]
    target_dt = datetime.datetime(year, month, 1)

    # Scan for existing month entry and Gesamt row
    gesamt_row = None
    last_data_row = 4
    existing_row = None

    for row in range(5, 100):
        val = ws.cell(row=row, column=1).value
        if isinstance(val, str) and val.strip() == "Gesamt":
            gesamt_row = row
            break
        if isinstance(val, datetime.datetime):
            last_data_row = row
            if val.year == year and val.month == month:
                existing_row = row

    new_row = existing_row if existing_row else last_data_row + 1

    # Write month entry
    a = ws.cell(row=new_row, column=1, value=target_dt)
    a.number_format = "mmm-yy"
    a.border = THIN_BORDER
    b = ws.cell(row=new_row, column=2, value=f"='{month_tab}'!D{ub_cell_row}")
    b.number_format = "0.00"
    b.border = THIN_BORDER

    # Move/update Gesamt row (2 rows below last data)
    new_gesamt_row = new_row + 2
    if gesamt_row and gesamt_row != new_gesamt_row:
        ws.cell(row=gesamt_row, column=1).value = None
        ws.cell(row=gesamt_row, column=2).value = None

    ws.cell(row=new_gesamt_row, column=1, value="Gesamt").font = FONT_10B
    ws.cell(row=new_gesamt_row, column=1).number_format = "mmm-yy"
    g = ws.cell(row=new_gesamt_row, column=2, value=f"=SUM(B5:B{new_row})")
    g.number_format = "0.00"
    g.border = THIN_BORDER

    print(f"  Überstunden tab: row {new_row} → ='{month_tab}'!D{ub_cell_row}")


def main():
    parser = argparse.ArgumentParser(description="Fill Stundenzettel from clawd tracker")
    parser.add_argument("--month", help="YYYY-MM (default: last complete month)")
    parser.add_argument("--dry-run", action="store_true", help="Preview only")
    args = parser.parse_args()

    # Determine target month
    if args.month:
        parts = args.month.split("-")
        year, month = int(parts[0]), int(parts[1])
    else:
        today = datetime.date.today()
        first = today.replace(day=1)
        prev = first - datetime.timedelta(days=1)
        year, month = prev.year, prev.month

    tname = tab_name(year, month)
    print(f"Target: {year}-{month:02d} (tab: {tname})")

    # Parse markdown
    if not CLAWD_MD.exists():
        print(f"ERROR: {CLAWD_MD} not found")
        print("  Did you run 'git pull' in clawd/?")
        sys.exit(1)

    entries = parse_month_entries(CLAWD_MD.read_text(), year, month)

    if not entries:
        print("WARNING: No completed time entries found for this month")

    total_sessions = sum(len(v) for v in entries.values())
    print(f"Entries: {total_sessions} session(s) across {len(entries)} day(s)")
    for d in sorted(entries):
        for e in entries[d]:
            note = f"  ({e['notes']})" if e["notes"] else ""
            print(f"  {d}  {e['start'].strftime('%H:%M')}-{e['end'].strftime('%H:%M')}{note}")

    if args.dry_run:
        print("\n[DRY RUN] No file written.")
        return

    # Open XLSX
    xlsx = find_latest_xlsx()
    print(f"\nBase file: {xlsx}")
    wb = openpyxl.load_workbook(xlsx)

    # Create month tab + update Überstunden
    sheet_name, ub_cell_row = create_month_sheet(wb, year, month, entries)
    update_ueberstunden_tab(wb, sheet_name, year, month, ub_cell_row)

    # Save with new name
    new_file = xlsx.parent / f"Stundenzettel-{tname}-Merlin.xlsx"
    wb.save(new_file)
    print(f"\nSaved: {new_file}")
    if new_file != xlsx:
        print(f"Previous file kept: {xlsx.name}")


if __name__ == "__main__":
    main()
